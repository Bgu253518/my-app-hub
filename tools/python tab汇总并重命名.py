# -*- coding: utf-8 -*-
import os
import sys
import tempfile
import subprocess
import pandas as pd
import PySimpleGUI as sg
from openpyxl import load_workbook, Workbook

# ==================== 自动安装依赖 ====================
def install_package(package):
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    except:
        pass

try:
    import pandas
    import PySimpleGUI
    from openpyxl import load_workbook
except:
    print("首次运行，自动安装依赖...")
    install_package("pandas")
    install_package("PySimpleGUI")
    install_package("openpyxl")

# ==================== 获取所有 Sheet ====================
def get_visible_sheets(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        wb.close()
        return sheets
    except:
        return []

# ==================== 安全表名 ====================
def safe_sheet_name(name, max_len=31):
    invalid = r'\/?*[]'
    for c in invalid:
        name = name.replace(c, "")
    return name[:max_len].strip()

def get_unique_name(wb, base):
    name = safe_sheet_name(base)
    counter = 1
    while name in wb.sheetnames:
        name = safe_sheet_name(f"{base[:27]}-{counter}")
        counter += 1
    return name

# ==================== 主逻辑 ====================
def main():
    sg.theme("SystemDefault")

    layout = [
        [sg.Text("📊 Excel 多文件智能合并（自动配置表 + 自动重命名）", font=("", 11))],
        [sg.Text("选择要合并的Excel文件：")],
        [sg.Input(key="FILES", size=(60, 1)), sg.FilesBrowse(file_types=(("Excel 文件", "*.xlsx;*.xlsm;*.xls"),))],
        [sg.Button("开始处理", size=(25, 1))],
        [sg.Multiline(size=(75, 15), key="LOG", autoscroll=True, disabled=True)],
    ]

    window = sg.Window("Excel 智能合并工具（无配置文件版）", layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break

        if event == "开始处理":
            file_str = values["FILES"].strip()
            if not file_str:
                sg.popup_error("请先选择文件！")
                continue

            files = file_str.split(";")
            files = [f.strip() for f in files if f.strip()]

            log = lambda m: window["LOG"].print(m)
            log("🚀 正在扫描所有工作表...")

            # 收集所有 sheet
            rows = []
            for f in files:
                name_only = os.path.splitext(os.path.basename(f))[0]
                sheets = get_visible_sheets(f)
                for s in sheets:
                    display = f"{name_only}-{s}"
                    rows.append([display, display, "Y", f, s])

            if not rows:
                log("❌ 未找到任何工作表")
                continue

            log(f"✅ 共找到 {len(rows)} 个工作表")

            # ==================== 生成临时配置表 ====================
            with tempfile.NamedTemporaryFile(mode='w+b', suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name

            df = pd.DataFrame(rows, columns=[
                "原名称", "新Sheet名称", "是否抓取(Y/N)", "源文件路径", "源Sheet名"
            ])
            df.to_excel(tmp_path, index=False, sheet_name="配置")

            log("📋 正在打开配置表，请修改后【保存并关闭Excel】...")
            os.startfile(tmp_path)
            sg.popup("提示", "修改完配置表后\n请保存并关闭 Excel\n然后点击确定继续")

            # ==================== 读取用户修改后的配置 ====================
            try:
                df_edit = pd.read_excel(tmp_path, sheet_name="配置")
            except:
                log("❌ 无法读取配置表，请关闭 Excel")
                continue

            # ==================== 选择输出路径 ====================
            output = sg.popup_get_file("保存合并结果", save_as=True, file_types=(("Excel 文件", "*.xlsx"),))
            if not output:
                log("❌ 已取消")
                continue

            # ==================== 开始合并 ====================
            log("🔨 开始合并并重命名工作表...")
            wb_final = Workbook()
            wb_final.remove(wb_final.active)

            total = 0
            for _, r in df_edit.iterrows():
                try:
                    if str(r["是否抓取(Y/N)"]).strip().upper() != "Y":
                        continue

                    src_file = str(r["源文件路径"])
                    src_sheet = str(r["源Sheet名"])
                    new_name = str(r["新Sheet名称"]).strip()

                    wb_src = load_workbook(src_file, data_only=True)
                    if src_sheet not in wb_src.sheetnames:
                        log(f"⚠️ 跳过：{src_sheet}")
                        continue

                    ws_src = wb_src[src_sheet]
                    final_name = get_unique_name(wb_final, new_name)
                    ws_new = wb_final.create_sheet(title=final_name)

                    for row in ws_src.iter_rows():
                        for cell in row:
                            ws_new[cell.coordinate].value = cell.value

                    wb_src.close()
                    total += 1
                    log(f"✅ {src_sheet} → {final_name}")
                except Exception as e:
                    log(f"❌ 失败：{str(e)}")

            wb_final.save(output)
            wb_final.close()
            os.unlink(tmp_path)

            log(f"\n🎉 合并完成！共 {total} 个工作表")
            log(f"📁 已保存到：{output}")
            os.startfile(os.path.dirname(output))

    window.close()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sg.popup_error("程序错误", str(e))
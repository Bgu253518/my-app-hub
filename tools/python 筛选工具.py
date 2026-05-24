import os
import threading
import time
import PySimpleGUI as sg
import pandas as pd
import numpy as np

# ==================== 界面设计 ====================
def create_main_window():
    sg.theme('SystemDefault')

    manual_col = sg.Column([
        [sg.Text('粘贴识别码（每行一个）：')],
        [sg.Multiline(size=(40, 8), key='code_col1')],
        [sg.Text('第二列（多条件时使用，行数需一致）：')],
        [sg.Multiline(size=(40, 4), key='code_col2', disabled=True)],
        [sg.Button('检查识别码配对', key='check_pair')]
    ], key='manual_panel', visible=True)

    file_col = sg.Column([
        [sg.Text('源文件（唯一识别码汇总）'), sg.Input(key='src_file', size=(30,1)), sg.FileBrowse('选择文件', file_types=(('Excel', '*.xlsx;*.xlsm'),))],
    ], key='file_panel', visible=False)

    tab1_layout = [
        [sg.Radio('手动输入识别码', 'src_mode', key='manual', default=True, enable_events=True),
         sg.Radio('从 Excel 文件导入', 'src_mode', key='from_file', enable_events=True),
         sg.Button('生成标准化模板 Excel', key='gen_template')],
        [manual_col],
        [file_col]
    ]

    tab2_layout = [
        [sg.Frame('目标文件', [
            [sg.Listbox(values=[], size=(60, 6), key='target_files', select_mode='extended')],
            [sg.Button('添加文件', key='add_files'), sg.Button('清空文件', key='clear_files')]
        ])],
        [sg.Frame('筛选与匹配设置', [
            [sg.Text('目标筛选列（如 E）'), sg.Input(key='tgt_col1', size=4, default_text='E'),
             sg.Text('第二列（如 F）'), sg.Input(key='tgt_col2', size=4, disabled=True, default_text='F'),
             sg.Text('金额列（如 K）'), sg.Input(key='amount_col', size=4, default_text='K')],
            [sg.Checkbox('多条件筛选（两列组合）', key='multi', enable_events=True),
             sg.Checkbox('精确匹配（推荐）', key='exact', default=True)]
        ])],
        [sg.Frame('输出设置', [
            [sg.Checkbox('合并所有文件结果到一张总表（自动拆分）', key='merge_all', default=True)],
            [sg.Checkbox('每个目标文件单独生成汇总', key='per_file', default=False)],
            [sg.Text('输出文件夹（留空则保存在目标文件同目录）')],
            [sg.Input(key='output_dir', size=(50,1)), sg.FolderBrowse('选择文件夹', key='folder_browse')],
            [sg.Text('每文件最大行数（万行）'), sg.Input(key='max_rows', size=5, default_text='50')]
        ])]
    ]

    bottom_layout = [
        [sg.Button('▶ 开始筛选', size=20), sg.Button('退出')],
        [sg.Multiline(size=(100, 12), key='log', autoscroll=True, disabled=True, background_color='white')],
        [sg.Text('', key='status')]
    ]

    tab_group = sg.TabGroup([
        [sg.Tab('识别码来源', tab1_layout, key='tab1'),
         sg.Tab('目标文件与设置', tab2_layout, key='tab2')]
    ])

    layout = [[tab_group], [sg.Column(bottom_layout)]]
    return sg.Window('智能筛选汇总工具（合并总表+自动拆分）', layout, size=(850, 750), finalize=True)

# ==================== 生成标准化模板（主线程调用） ====================
def generate_template():
    path = sg.popup_get_file(
        '保存模板到...',
        save_as=True,
        default_extension='.xlsx',
        file_types=(('Excel 工作簿', '*.xlsx'),)
    )
    if not path:
        return
    if not path.lower().endswith('.xlsx'):
        path += '.xlsx'

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '唯一识别码列1'
        ws['B1'] = '唯一识别码列2'
        ws['A1'].font = openpyxl.styles.Font(bold=True)
        ws['B1'].font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        wb.save(path)
        wb.close()
        try:
            os.startfile(path)
        except:
            pass
        sg.popup_ok(f'模板已生成：\n{path}')
    except Exception as e:
        sg.popup_error(f'生成模板失败：{e}')

# ==================== 列字母转索引 ====================
def col_letter_to_index(letter):
    letter = letter.upper()
    col = 0
    for ch in letter:
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return col - 1

# ==================== 向量化文本标准化 ====================
def normalize_series_vectorized(ser):
    s = ser.astype(str).str.strip().replace(r'\s+', ' ', regex=True)
    date_mask = s.str.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}')
    if date_mask.any():
        dates = pd.to_datetime(s[date_mask], errors='coerce')
        formatted = dates.dt.year.astype(str) + '/' + dates.dt.month.astype(str) + '/' + dates.dt.day.astype(str)
        s = s.copy()
        s[date_mask] = formatted[date_mask]
    return s

def build_keys_vectorized(df, col1_idx, col2_idx, multi, exact_match):
    col1 = normalize_series_vectorized(df.iloc[:, col1_idx])
    if multi:
        col2 = normalize_series_vectorized(df.iloc[:, col2_idx])
        if exact_match:
            keys = col1 + "||" + col2
        else:
            keys = col1.str.upper().str.replace(' ', '') + "||" + col2.str.upper().str.replace(' ', '')
    else:
        if exact_match:
            keys = col1
        else:
            keys = col1.str.upper().str.replace(' ', '')
    return keys

# ==================== 拆分写入独立文件 ====================
def save_split_files(base_path, summary_df, check_df, max_rows_per_file, log):
    total_rows = len(summary_df)
    num_parts = max(1, (total_rows + max_rows_per_file - 1) // max_rows_per_file)
    for part in range(num_parts):
        start_idx = part * max_rows_per_file
        end_idx = min(start_idx + max_rows_per_file, total_rows)
        part_summary = summary_df.iloc[start_idx:end_idx].copy()
        if num_parts > 1:
            file_path = f"{base_path}_part{part + 1}.xlsx"
        else:
            file_path = f"{base_path}.xlsx"
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            part_summary.to_excel(writer, sheet_name='筛选汇总', index=False)
            check_df.to_excel(writer, sheet_name='金额验证', index=False)
        log(f'  已生成：{file_path} ({len(part_summary)} 行)')

# ==================== 核心筛选 ====================
def run_filter(src_mode, manual_data, src_file, amount_col, target_paths,
               tgt_col1, tgt_col2, multi, exact_match,
               per_file, merge_all, output_dir, max_rows_str, log, set_status):
    t_start = time.time()
    try:
        max_rows_per_file = int(float(max_rows_str.strip()) * 10000) if max_rows_str.strip() else 500000
    except:
        max_rows_per_file = 500000

    if not per_file and not merge_all:
        log('错误：至少选择一种输出方式。')
        return

    log('正在加载识别码...')
    # 收集识别码
    if src_mode == 'manual':
        lines1 = [x.strip() for x in manual_data['col1'].strip().splitlines() if x.strip()]
        lines2 = [x.strip() for x in manual_data['col2'].strip().splitlines() if x.strip()] if multi else []
        if multi and len(lines1) != len(lines2):
            log('错误：两列行数不一致！')
            return
        df_criteria = pd.DataFrame({'c1': lines1, 'c2': lines2 if multi else None})
        if multi:
            keys = build_keys_vectorized(df_criteria, 0, 1, True, exact_match)
        else:
            keys = normalize_series_vectorized(df_criteria['c1'])
            if not exact_match:
                keys = keys.str.upper().str.replace(' ', '')
    else:
        log(f'读取源文件：{src_file}')
        try:
            src_df = pd.read_excel(src_file, engine='calamine', dtype=str)
        except:
            src_df = pd.read_excel(src_file, dtype=str)
        if src_df.shape[1] < 1:
            log('源文件无数据。')
            return
        if multi and src_df.shape[1] < 2:
            log('源文件列数不足，无法获取第二列识别码。')
            return
        keys = build_keys_vectorized(src_df, 0, 1 if multi else None, multi, exact_match)

    criteria = set(keys.dropna().unique())
    log(f'识别码数量：{len(criteria)} (耗时 {time.time()-t_start:.1f}秒)')
    if not criteria:
        log('识别码为空，终止。')
        return

    tgt_idx1 = col_letter_to_index(tgt_col1)
    tgt_idx2 = col_letter_to_index(tgt_col2) if multi and tgt_col2 else None
    amount_idx = col_letter_to_index(amount_col)

    global_matched_parts = []
    global_total_amount = 0.0
    total_rows_all = 0

    for idx, tgt_file in enumerate(target_paths, 1):
        if not os.path.isfile(tgt_file):
            log(f'跳过无效路径：{tgt_file}')
            continue
        file_basename = os.path.basename(tgt_file)
        set_status(f'处理中：{idx}/{len(target_paths)} - {file_basename}')
        log(f'\n处理 [{idx}/{len(target_paths)}]：{file_basename}')
        t_file_start = time.time()

        try:
            sheets_dict = pd.read_excel(tgt_file, sheet_name=None, engine='calamine', dtype=str)
        except:
            sheets_dict = pd.read_excel(tgt_file, sheet_name=None, dtype=str)

        file_matched_parts = []
        file_total_rows = 0
        file_total_amount = 0.0
        sheet_amount_info = []

        for sheet_name, df in sheets_dict.items():
            if df.empty or tgt_idx1 >= df.shape[1]:
                continue
            if multi and tgt_idx2 is not None and tgt_idx2 >= df.shape[1]:
                continue

            keys = build_keys_vectorized(df, tgt_idx1, tgt_idx2, multi, exact_match)
            match_mask = keys.isin(criteria)
            n_match = match_mask.sum()
            if n_match == 0:
                continue

            sheet_amount = 0.0
            if amount_idx < df.shape[1]:
                amt_series = pd.to_numeric(df.iloc[:, amount_idx], errors='coerce')
                sheet_amount = amt_series[match_mask].sum()

            sheet_amount_info.append((sheet_name, n_match, sheet_amount))
            file_total_rows += n_match
            file_total_amount += sheet_amount

            matched_df = df.loc[match_mask].copy()
            matched_df.insert(0, '来源工作表', sheet_name)
            matched_df.insert(1, '来源文件', file_basename)
            file_matched_parts.append(matched_df)

            log(f'  [{sheet_name}] 匹配 {n_match} 行，金额 {sheet_amount:,.2f}')

        if file_total_rows > 0:
            total_rows_all += file_total_rows
            global_total_amount += file_total_amount

            if per_file:
                file_summary = pd.concat(file_matched_parts, ignore_index=True)
                file_check_df = pd.DataFrame(sheet_amount_info,
                                             columns=['工作表名称', '筛选行数', f'{amount_col}列合计金额'])
                file_total_row = pd.DataFrame([['合计', file_total_rows, file_total_amount]],
                                              columns=file_check_df.columns)
                file_check_df = pd.concat([file_check_df, file_total_row], ignore_index=True)

                save_dir = output_dir if output_dir and os.path.isdir(output_dir) else os.path.dirname(tgt_file)
                base_name = os.path.join(save_dir, f"单文件汇总_{file_basename.replace('.xlsx','')}")
                save_split_files(base_name, file_summary, file_check_df, max_rows_per_file, log)

            global_matched_parts.extend(file_matched_parts)
            log(f'  本文件耗时 {time.time()-t_file_start:.1f}秒')
        else:
            log('  未找到匹配数据，跳过。')

    if total_rows_all == 0:
        log('所有文件均未匹配到数据。')
        return

    if merge_all:
        log('\n=== 生成合并总表 ===')
        if not global_matched_parts:
            log('无数据可合并。')
            return
        mega_summary = pd.concat(global_matched_parts, ignore_index=True)

        if '来源文件' in mega_summary.columns:
            file_groups = mega_summary.groupby('来源文件')
            file_stats = []
            actual_amount_col = amount_idx + 2
            for fname, grp in file_groups:
                amt = 0.0
                if actual_amount_col < mega_summary.shape[1]:
                    amt_series = pd.to_numeric(grp.iloc[:, actual_amount_col], errors='coerce')
                    amt = amt_series.sum()
                file_stats.append([fname, len(grp), amt])
            check_df = pd.DataFrame(file_stats, columns=['来源文件', '筛选行数', f'{amount_col}列合计金额'])
        else:
            check_df = pd.DataFrame(columns=['来源文件', '筛选行数', f'{amount_col}列合计金额'])

        total_row = pd.DataFrame([['合计', total_rows_all, global_total_amount]],
                                 columns=check_df.columns)
        check_df = pd.concat([check_df, total_row], ignore_index=True)

        save_dir = output_dir if output_dir and os.path.isdir(output_dir) else os.path.dirname(target_paths[0])
        merge_base = os.path.join(save_dir, "合并汇总")
        save_split_files(merge_base, mega_summary, check_df, max_rows_per_file, log)
        log('✅ 合并总表生成完毕。')

    elapsed = time.time() - t_start
    status_msg = f'全部完成！总筛选行数 {total_rows_all}，总金额 {global_total_amount:,.2f}，总耗时 {elapsed:.1f}秒'
    log(f'\n{status_msg}')
    set_status(status_msg)

# ==================== 检查配对 ====================
def check_pairing(manual_data, multi, window):
    col1 = [x.strip() for x in manual_data['col1'].strip().splitlines() if x.strip()]
    col2 = [x.strip() for x in manual_data['col2'].strip().splitlines() if x.strip()] if multi else []
    if multi:
        max_show = min(5, len(col1), len(col2))
        lines = []
        for i in range(max_show):
            lines.append(f"第{i+1}行: [{col1[i]}]  <-->  [{col2[i]}]")
        window['log'].print("\n".join(lines))
        if len(col1) != len(col2):
            window['log'].print(f"⚠️ 警告：两列行数不一致（第一列{len(col1)}行，第二列{len(col2)}行）")
    else:
        window['log'].print("当前为单条件模式，无需配对检查。")

# ==================== 主循环 ====================
def main():
    window = create_main_window()
    target_paths = []

    def update_ui():
        manual = window['manual'].get()
        multi = window['multi'].get()
        window['manual_panel'].update(visible=manual)
        window['file_panel'].update(visible=not manual)
        window['code_col2'].update(disabled=not (manual and multi))
        window['tgt_col2'].update(disabled=not multi)

    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, '退出'): break

        if event in ('manual', 'from_file', 'multi'):
            update_ui()

        # 修复点：直接调用，不在子线程中
        if event == 'gen_template':
            generate_template()

        if event == 'check_pair':
            manual_data = {'col1': values['code_col1'], 'col2': values['code_col2']}
            check_pairing(manual_data, values['multi'], window)

        if event == 'add_files':
            files_str = sg.popup_get_file('选择目标Excel文件', multiple_files=True,
                                          file_types=(('Excel', '*.xlsx;*.xlsm'),))
            if files_str:
                new_files = [f.strip() for f in files_str.split(';') if f.strip()]
                target_paths.extend(new_files)
                seen = set()
                unique = []
                for p in target_paths:
                    if p not in seen:
                        seen.add(p); unique.append(p)
                target_paths = unique
                window['target_files'].update(target_paths)

        if event == 'clear_files':
            target_paths.clear()
            window['target_files'].update([])

        if event == 'folder_browse':
            folder = values['folder_browse']
            if folder:
                window['output_dir'].update(folder)

        if event == '▶ 开始筛选':
            src_mode = 'manual' if values['manual'] else 'file'
            manual_data = {'col1': values['code_col1'], 'col2': values['code_col2']}
            src_file = values['src_file']
            amount_col = values['amount_col'].strip().upper()
            tgt_col1 = values['tgt_col1'].strip().upper()
            tgt_col2 = values['tgt_col2'].strip().upper() if values['multi'] else None
            exact = values['exact']
            multi = values['multi']
            per_file = values['per_file']
            merge_all = values['merge_all']
            output_dir = values['output_dir']
            max_rows_str = values['max_rows']

            if src_mode == 'file' and not src_file:
                sg.popup_error('请选择源文件！'); continue
            if src_mode == 'manual' and not manual_data['col1'].strip():
                sg.popup_error('请粘贴至少一个识别码！'); continue
            if not target_paths:
                sg.popup_error('请添加至少一个目标文件！'); continue
            if not amount_col or not tgt_col1:
                sg.popup_error('请填写金额列和目标筛选列！'); continue
            if not per_file and not merge_all:
                sg.popup_error('请至少选择一种输出方式（每个文件单独汇总 或 合并总表）！'); continue

            window['log'].update('')
            log = lambda msg: window['log'].print(msg)
            set_status = lambda msg: window['status'].update(msg)

            threading.Thread(target=run_filter,
                             args=(src_mode, manual_data, src_file, amount_col,
                                   target_paths.copy(), tgt_col1, tgt_col2,
                                   multi, exact, per_file, merge_all,
                                   output_dir, max_rows_str, log, set_status),
                             daemon=True).start()

    window.close()

if __name__ == '__main__':
    main()